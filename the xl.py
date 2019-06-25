from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import csv
import pandas
import time
import openpyxl
import xlsxwriter
import matplotlib.pyplot as plt

browser = webdriver.Chrome("D:\Downloads\chromedriver")
h = []
df = pandas.read_csv(r"C:\Users\PRITIKA\Documents\Cities1.csv")
wb = openpyxl.Workbook()
sheet = wb.active
wb.create_sheet(index=1, title="Processing Time")


for i in range(10):
    data = df['Name'][i]
    h.append(data)


def open_function():
    browser.get("https://www.magicbricks.com")
    browser.implicitly_wait(5)
    n = []
    for i in range(len(h)):
        s = 'https://www.magicbricks.com/property-for-sale/residential-real-estate?proptype=Multistorey-Apartment,Builder-Floor-Apartment,Penthouse,Studio-Apartment,Residential-House,Villa&cityName='
        s = s + h[i]
        start=time.time()
        browser.get(s)
        end = time.time()
        p = (end - start)
        n.append(p)
        c1 = sheet.cell(row=i+1, column=1)
        c1.value =p
        wb.save(r"C:\Users\PRITIKA\Documents\LoadingTime.xls")
    print(n)


    plt.xlabel('x - axis')
    plt.ylabel('y - axis')
    plt.title('Graph')
    plt.plot(h,n)
    plt.show()

open_function()